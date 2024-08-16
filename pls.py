import re
import openpyxl
wb=openpyxl.load_workbook('data.xlsx')
sheet=wb.get_sheet_by_name('Sheet1')
for i in range(2,47241,1):
    x=sheet.cell(row=i,column=1).value
    y=re.findall('\d+',x)
    if not y:
        y.extend(['a'])
    z=re.findall('[A-Za-z]',x)
    z=''.join(z)

    if z=='EA':
        text='1EA'
    elif z=='BX':
        text='1BX'+y[0]
    elif z=='BG':
        text='1BG'+y[0]
    elif z=='PK':
        text='1PK'+y[0]
    elif z=='CA':
        text='1CS'+y[0]

    sheet.cell(row=i,column=2).value=text


    


wb.save('data_copy.xlsx')
