from bs4 import BeautifulSoup
import requests
import openpyxl
source=requests.get('https://www.vitalitymedical.com/non-adhesive-tape.html').text
soup=BeautifulSoup(source,'lxml')

wb=openpyxl.load_workbook('tapee.xlsx')
sheet=wb.get_sheet_by_name('Sheet1')

r=1
for product in soup.find_all('li', class_="item last"):

    
    title=product.find('h2', class_="product-name").text
    
    link=product.find('a')['href']

    manuname=product.find('span', class_="manu-small").text
    

    newsource=requests.get(link).text
    new_soup=BeautifulSoup(newsource,'lxml')

    x=new_soup.find('table', id="super-product-table")
    table_rows=x.find_all('tr')

    rr=r
    x=1

    for tr in table_rows:

        sheet.cell(row=rr,column=1).value=title
        sheet.cell(row=rr,column=2).value=manuname
        
        pid=tr.find('td', class_="manu-sku")
        if pid is not None:
            sheet.cell(row=rr,column=3).value=pid.text
            
        des=tr.find('td', class_="name")
        if des is not None:
            sheet.cell(row=rr,column=4).value=des.text

        size=tr.find('td', class_="uom")
        if size is not None:
            sheet.cell(row=rr,column=5).value=size.text
     
        
        td=tr.find_all('td', class_="price-col")
        for i in td:
            price=i.find('p', class_="special-price")
            if price is not None:
                pricex=price.find('span',class_="price")
                sheet.cell(row=rr,column=6).value=pricex.text
            if price is None:
                pricey=i.find('span', class_="special-price")
                pricey=pricey.find('span', class_="price")
                sheet.cell(row=rr,column=6).value=pricey.text
                
        if x!=1:        
            rr+=1
            r+=1
        x+=1


wb.save("tapee.xlsx")

    

                           
