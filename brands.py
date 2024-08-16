from bs4 import BeautifulSoup
import openpyxl
import requests

source=requests.get('https://www.vitalitymedical.com/brands/').text
soup=BeautifulSoup(source,'lxml')

r=1
wb=openpyxl.load_workbook('tapee.xlsx')
sheet=wb.get_sheet_by_name('Sheet1')


for br in soup.find_all('li', class_="alph-sub"):
    for li in br.find_all('a'):
        wb.save("tapee.xlsx")
        brlink="https://www.vitalitymedical.com/" + li['href'] + "?manufacturer=3956&limit=20"
        neww_source=requests.get(brlink).text
        neww_soup=BeautifulSoup(neww_source,'lxml')
        
        page=neww_soup.find('div', class_="count-container")
        if page.find('p', class_="amount amount--has-pages") is not None:
            y=page.find('p', class_="amount amount--has-pages").text.split()
            q=(int(y[2]))

            if q%20==0:
                p=int(q/20)
            else:
                p=int( (q/20)+1)

            for i in range(0,p,1):
                for product in neww_soup.find_all('li', class_="item last"):   
                    title=product.find('h2', class_="product-name").text
                    print(title)
                    link=product.find('a')['href']
                    manuname=product.find('span', class_="manu-small").text
                    
                    newsource=requests.get(link).text
                    new_soup=BeautifulSoup(newsource,'lxml')

                    x=new_soup.find('table', id="super-product-table")
                    table_rows=x.find_all('tr')

                    rr=r
                    z=1

                    for tr in table_rows:

                        sheet.cell(row=rr,column=1).value=title
                        sheet.cell(row=rr,column=2).value=manuname
                        
                        pid=tr.find('td', class_="manu-sku")
                        if pid is not None:
                            sheet.cell(row=rr,column=3).value=pid.text
                            
                        des=tr.find('td', class_="name")
                        if des is not None:
                            sheet.cell(row=rr,column=4).value=des.text

                        size=tr.find('td', class_="uom")
                        if size is not None:
                            sheet.cell(row=rr,column=5).value=size.text
                     
                        
                        td=tr.find_all('td', class_="price-col")
                        for i in td:
                            price=i.find('p', class_="special-price")
                            if price is not None:
                                pricex=price.find('span',class_="price")
                                sheet.cell(row=rr,column=6).value=pricex.text
                            if price is None:
                                pricey=i.find('span', class_="special-price")
                                pricey=pricey.find('span', class_="price")
                                sheet.cell(row=rr,column=6).value=pricey.text                        

                        if z!=1:        
                            rr+=1
                            r+=1
                        z+=1
                        
                tool=neww_soup.find('div', class_="toolbar")
                pg=tool.find('div', class_="pages")
                lin=pg.find('a', class_="next i-next")
                if lin is not None:
                    linn="https://www.vitalitymedical.com/" + lin['href']
                    neww_source=requests.get(linn).text
                    neww_soup=BeautifulSoup(neww_source,'lxml')


        else:    
            for product in neww_soup.find_all('li', class_="item last"):   
                title=product.find('h2', class_="product-name").text
                print(title)
                link=product.find('a')['href']
                manuname=product.find('span', class_="manu-small").text

                newsource=requests.get(link).text
                new_soup=BeautifulSoup(newsource,'lxml')

                x=new_soup.find('table', id="super-product-table")
                table_rows=x.find_all('tr')

                rr=r
                z=1

                for tr in table_rows:

                    sheet.cell(row=rr,column=1).value=title
                    sheet.cell(row=rr,column=2).value=manuname
                    
                    pid=tr.find('td', class_="manu-sku")
                    if pid is not None:
                        sheet.cell(row=rr,column=3).value=pid.text
                        
                    des=tr.find('td', class_="name")
                    if des is not None:
                        sheet.cell(row=rr,column=4).value=des.text

                    size=tr.find('td', class_="uom")
                    if size is not None:
                        sheet.cell(row=rr,column=5).value=size.text
                 
                    
                    td=tr.find_all('td', class_="price-col")
                    for i in td:
                        price=i.find('p', class_="special-price")
                        if price is not None:
                            pricex=price.find('span',class_="price")
                            sheet.cell(row=rr,column=6).value=pricex.text
                        if price is None:
                            pricey=i.find('span', class_="special-price")
                            pricey=pricey.find('span', class_="price")
                            sheet.cell(row=rr,column=6).value=pricey.text                        

                    if z!=1:        
                        rr+=1
                        r+=1
                    z+=1


wb.save("tapee.xlsx")
