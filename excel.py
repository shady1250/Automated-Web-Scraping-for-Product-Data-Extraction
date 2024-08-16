import re
import openpyxl

wb=openpyxl.load_workbook('fi.xlsx')
sheet=wb.get_sheet_by_name('Sheet1')
for i in range(1,1614,1):
    
    if sheet.cell(row=i,column=1).value is not None:
        x=sheet.cell(row=i,column=1).value
        if x=="Each":
            text='1EA'
        elif re.search("^1BX",x):
            text=x
        elif re.match("Count",x):
            text=x
        elif re.fullmatch("Case",x):
            text="1CS"
        else:
            z=x.split()
            if z[0]=="Case":
                text="1CS" + z[2]
            elif z[0]=="Pack":
                text="1PK" + z[2]
            elif z[0]=="Box":
                text="1BX" + z[2]
            elif z[1]=="Packs/Case":
                text="1CS" + z[2]
            else:
                text=x

        sheet.cell(row=i, column=2).value=text

wb.save('fi.xlsx')

